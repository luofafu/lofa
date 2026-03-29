from openpyxl import  load_workbook

workbook = load_workbook(filename='./ecommerce_data_50.xlsx')
print(workbook.sheetnames)
sheet = workbook.active
print(sheet)

print(sheet.dimensions)
cell = sheet['A1']
cell = cell.value
print(cell)

